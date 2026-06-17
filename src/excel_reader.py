"""Read bank rows from Excel."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.models import BankRow

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = [
    "bank_name",
    "bank_legal_name",
    "recipient_email",
    "cc_email",
    "bcc_email",
    "chair_full_name",
    "chair_short_dative",
    "gender",
    "greeting_name",
    "pdf_filename",
    "email_bank_name",
]


class ExcelReaderError(Exception):
    """Raised when Excel file cannot be read or is invalid."""


def read_banks(excel_path: Path) -> list[BankRow]:
    """Load bank rows from the first sheet of an Excel workbook."""
    if not excel_path.exists():
        raise ExcelReaderError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ExcelReaderError("Workbook has no active sheet")

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ExcelReaderError("Excel file is empty")

    header_idx = _find_header_row_index(all_rows)
    if header_idx is None:
        raise ExcelReaderError(
            f"Header row with columns {REQUIRED_COLUMNS[:3]}... not found in first 30 rows"
        )

    headers = [_normalize_header(h) for h in all_rows[header_idx]]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ExcelReaderError(f"Missing required columns: {', '.join(missing)}")

    col_index = {name: headers.index(name) for name in REQUIRED_COLUMNS}
    result: list[BankRow] = []
    excel_row_num = header_idx + 1

    for values in all_rows[header_idx + 1 :]:
        excel_row_num += 1
        if _is_empty_row(values):
            continue

        data: dict[str, Any] = {}
        for col_name, idx in col_index.items():
            if idx < len(values):
                data[col_name] = values[idx]
            else:
                data[col_name] = None

        result.append(BankRow.from_excel_dict(excel_row_num, data))

    logger.info("Loaded %d bank row(s) from %s (header at row %d)", len(result), excel_path, header_idx + 1)
    return result


def _find_header_row_index(rows: list[tuple[Any, ...]], max_scan: int = 30) -> int | None:
    """Find row index that contains all required column headers."""
    for idx, row in enumerate(rows[:max_scan]):
        headers = {_normalize_header(h) for h in row if h is not None}
        if all(col in headers for col in REQUIRED_COLUMNS):
            return idx
    return None


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)
