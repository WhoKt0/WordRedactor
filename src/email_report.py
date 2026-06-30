"""Email delivery report for a generation manifest."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

EMAIL_REPORT_COLUMNS = [
    "generation_id",
    "email_run_id",
    "mode",
    "excel_row",
    "bank_name",
    "bank_legal_name",
    "out_number",
    "recipient_email",
    "cc_email",
    "bcc_email",
    "subject",
    "pdf_path",
    "status",
    "attempts",
    "sent_at",
    "error_message",
]

EMAIL_STATUS_PENDING = "pending"
EMAIL_STATUS_DRY_RUN = "dry_run"
EMAIL_STATUS_SENT = "sent"
EMAIL_STATUS_FAILED = "failed"
EMAIL_STATUS_SKIPPED_PDF_MISSING = "skipped_pdf_missing"
EMAIL_STATUS_SKIPPED_INVALID_EMAIL = "skipped_invalid_email"
EMAIL_STATUS_TEST_SENT = "test_sent"
EMAIL_STATUS_TEST_FAILED = "test_failed"

EMAIL_TEST_REPORT_COLUMNS = [
    "generation_id",
    "email_run_id",
    "mode",
    "excel_row",
    "bank_name",
    "bank_legal_name",
    "out_number",
    "original_recipient_email",
    "test_recipient_email",
    "cc_email",
    "bcc_email",
    "subject",
    "pdf_path",
    "status",
    "attempts",
    "sent_at",
    "error_message",
]


def _safe_int(value: str, default: int = 0) -> int:
    if not value:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


@dataclass
class EmailReportRow:
    generation_id: int
    email_run_id: str
    mode: str
    excel_row: int
    bank_name: str
    bank_legal_name: str
    out_number: int
    recipient_email: str
    cc_email: str
    bcc_email: str
    subject: str
    pdf_path: str
    status: str = EMAIL_STATUS_PENDING
    attempts: int = 0
    sent_at: str = ""
    error_message: str = ""

    def resume_key(self) -> tuple[int, int, str, str]:
        return (
            self.generation_id,
            self.out_number,
            self.recipient_email.strip().lower(),
            self.pdf_path.replace("\\", "/"),
        )


@dataclass
class EmailReport:
    generation_id: int
    report_path: Path
    rows: list[EmailReportRow] = field(default_factory=list)


@dataclass
class EmailTestReportRow:
    generation_id: int
    email_run_id: str
    mode: str
    excel_row: int
    bank_name: str
    bank_legal_name: str
    out_number: int
    original_recipient_email: str
    test_recipient_email: str
    cc_email: str
    bcc_email: str
    subject: str
    pdf_path: str
    status: str = EMAIL_STATUS_TEST_FAILED
    attempts: int = 0
    sent_at: str = ""
    error_message: str = ""


@dataclass
class EmailTestReport:
    generation_id: int
    report_path: Path
    rows: list[EmailTestReportRow] = field(default_factory=list)


def email_report_path(reports_dir: Path, generation_id: int) -> Path:
    return reports_dir / f"ген_{generation_id}_email_report.xlsx"


def email_test_report_path(reports_dir: Path, generation_id: int) -> Path:
    return reports_dir / f"ген_{generation_id}_email_test_report.xlsx"


def write_email_report(report: EmailReport) -> Path:
    report.report_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "email_report"

    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(EMAIL_REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, email_row in enumerate(report.rows, start=2):
        values = [
            email_row.generation_id,
            email_row.email_run_id,
            email_row.mode,
            email_row.excel_row,
            email_row.bank_name,
            email_row.bank_legal_name,
            email_row.out_number,
            email_row.recipient_email,
            email_row.cc_email,
            email_row.bcc_email,
            email_row.subject,
            email_row.pdf_path,
            email_row.status,
            email_row.attempts,
            email_row.sent_at,
            email_row.error_message,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(report.report_path)
    return report.report_path


def read_email_report(path: Path) -> EmailReport | None:
    if not path.exists():
        return None

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
    column_index = {name: idx for idx, name in enumerate(header)}

    def _cell(row_values: tuple, name: str, default: str = "") -> str:
        idx = column_index.get(name)
        if idx is None:
            return default
        value = row_values[idx]
        if value is None:
            return default
        return str(value).strip()

    rows: list[EmailReportRow] = []
    generation_id = 0
    for row_values in rows_iter:
        if not row_values or row_values[0] is None:
            continue
        generation_id = _safe_int(_cell(row_values, "generation_id", "0"))
        rows.append(
            EmailReportRow(
                generation_id=generation_id,
                email_run_id=_cell(row_values, "email_run_id"),
                mode=_cell(row_values, "mode"),
                excel_row=_safe_int(_cell(row_values, "excel_row", "0")),
                bank_name=_cell(row_values, "bank_name"),
                bank_legal_name=_cell(row_values, "bank_legal_name"),
                out_number=_safe_int(_cell(row_values, "out_number", "0")),
                recipient_email=_cell(row_values, "recipient_email"),
                cc_email=_cell(row_values, "cc_email"),
                bcc_email=_cell(row_values, "bcc_email"),
                subject=_cell(row_values, "subject"),
                pdf_path=_cell(row_values, "pdf_path"),
                status=_cell(row_values, "status", EMAIL_STATUS_PENDING),
                attempts=_safe_int(_cell(row_values, "attempts", "0")),
                sent_at=_cell(row_values, "sent_at"),
                error_message=_cell(row_values, "error_message"),
            )
        )

    wb.close()
    if not rows:
        return None

    return EmailReport(
        generation_id=generation_id,
        report_path=path,
        rows=rows,
    )


def sent_resume_keys(report: EmailReport) -> set[tuple[int, int, str, str]]:
    return {
        row.resume_key()
        for row in report.rows
        if row.status == EMAIL_STATUS_SENT
    }


def write_email_test_report(report: EmailTestReport) -> Path:
    report.report_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "email_test_report"

    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(EMAIL_TEST_REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, email_row in enumerate(report.rows, start=2):
        values = [
            email_row.generation_id,
            email_row.email_run_id,
            email_row.mode,
            email_row.excel_row,
            email_row.bank_name,
            email_row.bank_legal_name,
            email_row.out_number,
            email_row.original_recipient_email,
            email_row.test_recipient_email,
            email_row.cc_email,
            email_row.bcc_email,
            email_row.subject,
            email_row.pdf_path,
            email_row.status,
            email_row.attempts,
            email_row.sent_at,
            email_row.error_message,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(report.report_path)
    return report.report_path
