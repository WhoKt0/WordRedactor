"""Manifest of successfully created documents per generation run."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

MANIFEST_COLUMNS = [
    "generation_id",
    "mode",
    "excel_row",
    "bank_name",
    "bank_legal_name",
    "email_bank_name",
    "chair_full_name",
    "greeting_word",
    "greeting_name_final",
    "template_used",
    "recipient_email",
    "cc_email",
    "bcc_email",
    "out_number",
    "letter_date",
    "docx_path",
    "pdf_path",
    "email_status",
    "created_at",
]


@dataclass
class ManifestRow:
    generation_id: int
    mode: str
    excel_row: int
    bank_name: str
    bank_legal_name: str
    email_bank_name: str
    chair_full_name: str
    greeting_word: str
    greeting_name_final: str
    template_used: str
    recipient_email: str
    cc_email: str
    bcc_email: str
    out_number: int
    letter_date: str
    docx_path: str
    pdf_path: str
    email_status: str
    created_at: str


def write_manifest_report(
    rows: list[ManifestRow],
    reports_dir: Path,
    generation_id: int,
    mode: str,
) -> Path | None:
    """Write manifest XLSX. Returns path if created, else None."""
    if not rows:
        return None

    reports_dir.mkdir(parents=True, exist_ok=True)
    report_path = reports_dir / f"ген_{generation_id}_manifest.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "manifest"

    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(MANIFEST_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, manifest_row in enumerate(rows, start=2):
        values = [
            manifest_row.generation_id,
            manifest_row.mode,
            manifest_row.excel_row,
            manifest_row.bank_name,
            manifest_row.bank_legal_name,
            manifest_row.email_bank_name,
            manifest_row.chair_full_name,
            manifest_row.greeting_word,
            manifest_row.greeting_name_final,
            manifest_row.template_used,
            manifest_row.recipient_email,
            manifest_row.cc_email,
            manifest_row.bcc_email,
            manifest_row.out_number,
            manifest_row.letter_date,
            manifest_row.docx_path,
            manifest_row.pdf_path,
            manifest_row.email_status,
            manifest_row.created_at,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(report_path)
    return report_path
