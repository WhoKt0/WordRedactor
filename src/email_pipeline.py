"""Email delivery pipeline over a final generation manifest."""

from __future__ import annotations

import logging
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from src.config import Settings
from src.email_report import (
    EMAIL_STATUS_DRY_RUN,
    EMAIL_STATUS_FAILED,
    EMAIL_STATUS_PENDING,
    EMAIL_STATUS_SENT,
    EMAIL_STATUS_SKIPPED_INVALID_EMAIL,
    EMAIL_STATUS_SKIPPED_PDF_MISSING,
    EMAIL_STATUS_TEST_FAILED,
    EMAIL_STATUS_TEST_SENT,
    EmailReport,
    EmailReportRow,
    EmailTestReport,
    EmailTestReportRow,
    email_report_path,
    email_test_report_path,
    read_email_report,
    sent_resume_keys,
    write_email_report,
    write_email_test_report,
)
from src.email_sender import (
    EmailSender,
    EmailSenderError,
    SMTP_HOST_LOOKS_LIKE_EMAIL,
    _split_emails,
    get_smtp_mode,
)
from src.output_utils import relative_project_path
from src.validators import is_valid_email

logger = logging.getLogger(__name__)

MANIFEST_FILENAME_PATTERN = re.compile(r"ген_(\d+)_manifest\.xlsx$", re.IGNORECASE)
PREVIEW_REPORTS_MARKER = "output/preview/reports"

SEND_CONFIRMATION_PROMPT = (
    "Вы точно хотите отправить реальные письма по этому manifest? Напишите SEND: "
)
TEST_EMAIL_PROMPT = "Введите тестовый email, на который отправить одно письмо: "

SMTP_NOT_CONFIGURED_MESSAGE = (
    "SMTP не настроен. Заполните .env: SMTP_HOST, SMTP_PORT, SMTP_USER, "
    "SMTP_PASSWORD, FROM_EMAIL."
)


class EmailPipelineError(Exception):
    """Raised when email pipeline cannot proceed."""


@dataclass
class ManifestEntry:
    generation_id: int
    mode: str
    excel_row: int
    bank_name: str
    bank_legal_name: str
    email_bank_name: str
    chair_full_name: str
    greeting_word: str
    greeting_name_final: str
    recipient_email: str
    cc_email: str
    bcc_email: str
    out_number: int
    letter_date: str
    pdf_path: str

    def resume_key(self) -> tuple[int, int, str, str]:
        return (
            self.generation_id,
            self.out_number,
            self.recipient_email.strip().lower(),
            self.pdf_path.replace("\\", "/"),
        )


@dataclass
class EmailPreflightSummary:
    manifest_path: Path
    generation_id: int
    total_rows: int
    sendable: int
    pdf_found: int
    pdf_missing: int
    invalid_email: int
    already_sent: int
    remaining: int
    out_number_min: int | None
    out_number_max: int | None
    first_recipients: list[str]


def _safe_int(value: str, default: int = 0) -> int:
    if not value:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def extract_generation_id(path: Path) -> int:
    match = MANIFEST_FILENAME_PATTERN.search(path.name)
    if not match:
        raise EmailPipelineError(f"Не удалось определить generation_id из имени файла: {path.name}")
    return int(match.group(1))


def _normalize_path_text(path_text: str) -> str:
    return path_text.replace("\\", "/").strip()


def is_preview_manifest_path(path: Path, project_root: Path) -> bool:
    try:
        rel = path.resolve().relative_to(project_root.resolve()).as_posix().lower()
    except ValueError:
        rel = path.as_posix().lower()
    return PREVIEW_REPORTS_MARKER in rel


def find_latest_final_manifest(reports_dir: Path) -> Path:
    candidates = sorted(
        reports_dir.glob("ген_*_manifest.xlsx"),
        key=lambda item: extract_generation_id(item),
    )
    if not candidates:
        raise EmailPipelineError(
            f"Manifest не найден в {reports_dir}. Сначала выполните final_generate.py."
        )
    return candidates[-1]


def resolve_manifest_path(
    *,
    project_root: Path,
    reports_dir: Path,
    manifest_arg: str | None,
) -> Path:
    if manifest_arg:
        manifest_path = Path(manifest_arg)
        if not manifest_path.is_absolute():
            manifest_path = (project_root / manifest_path).resolve()
    else:
        manifest_path = find_latest_final_manifest(reports_dir).resolve()

    if not manifest_path.exists():
        raise EmailPipelineError(f"Manifest не найден: {manifest_path}")

    if is_preview_manifest_path(manifest_path, project_root):
        raise EmailPipelineError(
            "Нельзя использовать preview manifest для рассылки. "
            "Укажите final manifest из output/reports/."
        )

    return manifest_path


def find_latest_preview_manifest(project_root: Path) -> Path:
    reports_dir = project_root / "output" / "preview" / "reports"
    candidates = sorted(
        reports_dir.glob("ген_*_manifest.xlsx"),
        key=lambda item: extract_generation_id(item),
    )
    if not candidates:
        raise EmailPipelineError(
            f"Preview manifest не найден в {reports_dir}. Сначала выполните preview.py."
        )
    return candidates[-1]


def resolve_preview_manifest_path(
    *,
    project_root: Path,
    manifest_arg: str | None,
) -> Path:
    if manifest_arg:
        manifest_path = Path(manifest_arg)
        if not manifest_path.is_absolute():
            manifest_path = (project_root / manifest_path).resolve()
    else:
        manifest_path = find_latest_preview_manifest(project_root).resolve()

    if not manifest_path.exists():
        raise EmailPipelineError(f"Manifest не найден: {manifest_path}")

    if not is_preview_manifest_path(manifest_path, project_root):
        raise EmailPipelineError(
            "preview_one_mail.py работает только с preview manifest "
            "из output/preview/reports/. Сначала выполните preview.py."
        )

    return manifest_path


def read_manifest(
    path: Path,
    project_root: Path,
    *,
    include_preview_rows: bool | None = None,
) -> list[ManifestEntry]:
    if include_preview_rows is None:
        include_preview_rows = is_preview_manifest_path(path, project_root)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
    column_index = {name: idx for idx, name in enumerate(header)}

    def _cell(row_values: tuple, name: str, default: str = "") -> str:
        idx = column_index.get(name)
        if idx is None:
            return default
        value = row_values[idx]
        if value is None:
            return default
        return str(value).strip()

    generation_id = extract_generation_id(path)
    entries: list[ManifestEntry] = []

    for row_values in rows_iter:
        if not row_values or row_values[0] is None:
            continue

        mode = _cell(row_values, "mode", "FINAL")
        if mode.upper() == "PREVIEW" and not include_preview_rows:
            continue

        bank_name = _cell(row_values, "bank_name")
        bank_legal_name = _cell(row_values, "bank_legal_name")
        email_bank_name = _cell(row_values, "email_bank_name") or bank_name
        pdf_path = _normalize_path_text(_cell(row_values, "pdf_path"))
        if pdf_path and not Path(pdf_path).is_absolute():
            pdf_path = relative_project_path(project_root, project_root / pdf_path)

        entries.append(
            ManifestEntry(
                generation_id=generation_id,
                mode=mode,
                excel_row=_safe_int(_cell(row_values, "excel_row", "0")),
                bank_name=bank_name,
                bank_legal_name=bank_legal_name,
                email_bank_name=email_bank_name,
                chair_full_name=_cell(row_values, "chair_full_name"),
                greeting_word=_cell(row_values, "greeting_word"),
                greeting_name_final=_cell(row_values, "greeting_name_final"),
                recipient_email=_cell(row_values, "recipient_email"),
                cc_email=_cell(row_values, "cc_email"),
                bcc_email=_cell(row_values, "bcc_email"),
                out_number=_safe_int(_cell(row_values, "out_number", "0")),
                letter_date=_cell(row_values, "letter_date"),
                pdf_path=pdf_path,
            )
        )

    wb.close()

    if not entries:
        raise EmailPipelineError(f"Manifest пуст: {path}")

    return entries


def validate_smtp_settings(settings: Settings, *, require_password: bool = True) -> None:
    env = settings.env
    missing: list[str] = []

    if not env.smtp_host:
        missing.append("SMTP_HOST")
    elif "@" in env.smtp_host:
        raise EmailPipelineError(SMTP_HOST_LOOKS_LIKE_EMAIL)

    if not env.smtp_port or env.smtp_port < 1 or env.smtp_port > 65535:
        missing.append("SMTP_PORT")

    if not env.smtp_user:
        missing.append("SMTP_USER")

    if require_password and not env.smtp_password:
        missing.append("SMTP_PASSWORD")

    if not env.from_email:
        missing.append("FROM_EMAIL")
    elif not is_valid_email(env.from_email):
        raise EmailPipelineError(f"Невалидный FROM_EMAIL: {env.from_email!r}")

    if missing:
        raise EmailPipelineError(SMTP_NOT_CONFIGURED_MESSAGE)

    if (
        env.smtp_user
        and env.from_email
        and env.smtp_user.strip().lower() != env.from_email.strip().lower()
    ):
        logger.warning(
            "SMTP_USER (%s) отличается от FROM_EMAIL (%s)",
            env.smtp_user,
            env.from_email,
        )
        print(
            f"Предупреждение: SMTP_USER ({env.smtp_user}) "
            f"отличается от FROM_EMAIL ({env.from_email})"
        )


def resolve_test_recipient_email(
    settings: Settings,
    *,
    override: str | None = None,
    prompt_if_missing: bool = True,
) -> str:
    """Return test recipient from --email, then TEST_MAIL in .env, else prompt."""
    if override and override.strip():
        email = override.strip()
    elif settings.env.test_mail.strip():
        email = settings.env.test_mail.strip()
        print(f"Тестовый email из .env (TEST_MAIL): {email}")
    elif prompt_if_missing:
        email = input(TEST_EMAIL_PROMPT).strip()
    else:
        raise EmailPipelineError(
            "Тестовый email не задан. Укажите TEST_MAIL в .env или флаг --email."
        )

    if not is_valid_email(email):
        raise EmailPipelineError(f"Невалидный тестовый email: {email!r}")
    return email


def validate_optional_emails(value: str) -> bool:
    if not value.strip():
        return True
    return all(is_valid_email(email) for email in _split_emails(value))


def validate_recipient_email(recipient_email: str) -> str | None:
    if not recipient_email.strip():
        return "recipient_email пустой"
    if not is_valid_email(recipient_email):
        return f"невалидный recipient_email: {recipient_email!r}"
    return None


def validate_cc_bcc(cc_email: str, bcc_email: str) -> str | None:
    if cc_email and not validate_optional_emails(cc_email):
        return f"невалидный cc_email: {cc_email!r}"
    if bcc_email and not validate_optional_emails(bcc_email):
        return f"невалидный bcc_email: {bcc_email!r}"
    return None


def resolve_pdf_path(project_root: Path, pdf_path_text: str) -> Path:
    pdf_path = Path(pdf_path_text)
    if not pdf_path.is_absolute():
        pdf_path = (project_root / pdf_path).resolve()
    return pdf_path


def validate_pdf(pdf_path: Path) -> str | None:
    if not pdf_path.exists():
        return f"PDF не найден: {pdf_path}"
    if pdf_path.stat().st_size <= 0:
        return f"PDF пустой: {pdf_path}"
    return None


def build_email_variables(entry: ManifestEntry) -> dict[str, str]:
    values = {
        "{{BANK_NAME}}": entry.bank_name,
        "{{EMAIL_BANK_NAME}}": entry.email_bank_name,
        "{{BANK_LEGAL_NAME}}": entry.bank_legal_name,
        "{{OUT_NUMBER}}": str(entry.out_number),
        "{{DATE}}": entry.letter_date,
        "{{GREETING_WORD}}": entry.greeting_word,
        "{{GREETING_NAME}}": entry.greeting_name_final,
        "{{CHAIR_FULL_NAME}}": entry.chair_full_name,
        "BANK_NAME": entry.bank_name,
        "EMAIL_BANK_NAME": entry.email_bank_name,
        "BANK_LEGAL_NAME": entry.bank_legal_name,
        "OUT_NUMBER": str(entry.out_number),
        "DATE": entry.letter_date,
        "GREETING_WORD": entry.greeting_word,
        "GREETING_NAME": entry.greeting_name_final,
        "CHAIR_FULL_NAME": entry.chair_full_name,
    }
    return values


def load_email_template(path: Path) -> str:
    if not path.exists():
        raise EmailPipelineError(f"Email template not found: {path}")
    return path.read_text(encoding="utf-8")


def make_email_run_id(generation_id: int) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{generation_id}_{timestamp}"


def build_preflight_summary(
    *,
    manifest_path: Path,
    entries: list[ManifestEntry],
    project_root: Path,
    already_sent_keys: set[tuple[int, int, str, str]],
) -> EmailPreflightSummary:
    pdf_found = 0
    pdf_missing = 0
    invalid_email = 0
    sendable = 0
    already_sent = 0
    first_recipients: list[str] = []
    out_numbers: list[int] = []

    for entry in entries:
        out_numbers.append(entry.out_number)

        pdf_error = validate_pdf(resolve_pdf_path(project_root, entry.pdf_path))
        if pdf_error:
            pdf_missing += 1
        else:
            pdf_found += 1

        email_error = validate_recipient_email(entry.recipient_email)
        cc_bcc_error = validate_cc_bcc(entry.cc_email, entry.bcc_email)
        if email_error or cc_bcc_error:
            invalid_email += 1

        if entry.resume_key() in already_sent_keys:
            already_sent += 1
            continue

        if pdf_error or email_error or cc_bcc_error:
            continue

        sendable += 1
        if len(first_recipients) < 5:
            first_recipients.append(entry.recipient_email)

    return EmailPreflightSummary(
        manifest_path=manifest_path,
        generation_id=extract_generation_id(manifest_path),
        total_rows=len(entries),
        sendable=sendable,
        pdf_found=pdf_found,
        pdf_missing=pdf_missing,
        invalid_email=invalid_email,
        already_sent=already_sent,
        remaining=sendable,
        out_number_min=min(out_numbers) if out_numbers else None,
        out_number_max=max(out_numbers) if out_numbers else None,
        first_recipients=first_recipients,
    )


def print_preflight_summary(
    summary: EmailPreflightSummary,
    *,
    project_root: Path,
    settings: Settings,
    delay_seconds: int,
) -> None:
    manifest_rel = relative_project_path(project_root, summary.manifest_path)
    print(f"Manifest: {manifest_rel}")
    print(f"Строк в manifest: {summary.total_rows}")
    print(f"PDF найдено: {summary.pdf_found}")
    print(f"PDF отсутствует: {summary.pdf_missing}")
    print(f"Невалидных email: {summary.invalid_email}")
    print(f"Писем к отправке: {summary.sendable}")
    if summary.already_sent:
        print(f"Уже отправлено ранее: {summary.already_sent}")
        print(f"Осталось отправить: {summary.remaining}")
    if summary.out_number_min is not None and summary.out_number_max is not None:
        print(f"Диапазон СВХИ: {summary.out_number_min}–{summary.out_number_max}")
    if summary.first_recipients:
        print("Первые получатели:")
        for recipient in summary.first_recipients:
            print(f"  - {recipient}")
    print(f"FROM_EMAIL: {settings.env.from_email or '(не задан)'}")
    print(f"FROM_NAME: {settings.env.from_name}")
    print(f"SMTP_HOST: {settings.env.smtp_host or '(не задан)'}")
    print(f"SMTP_PORT: {settings.env.smtp_port}")
    print(f"SMTP mode: {get_smtp_mode(settings.env.smtp_port)}")
    print(f"Delay between emails: {delay_seconds} seconds")


def _prepare_report_row(
    *,
    entry: ManifestEntry,
    email_run_id: str,
    subject: str,
    status: str,
    attempts: int = 0,
    sent_at: str = "",
    error_message: str = "",
) -> EmailReportRow:
    return EmailReportRow(
        generation_id=entry.generation_id,
        email_run_id=email_run_id,
        mode=entry.mode,
        excel_row=entry.excel_row,
        bank_name=entry.bank_name,
        bank_legal_name=entry.bank_legal_name,
        out_number=entry.out_number,
        recipient_email=entry.recipient_email,
        cc_email=entry.cc_email,
        bcc_email=entry.bcc_email,
        subject=subject,
        pdf_path=entry.pdf_path,
        status=status,
        attempts=attempts,
        sent_at=sent_at,
        error_message=error_message,
    )


def _render_subject_and_body(
    *,
    settings: Settings,
    email_sender: EmailSender,
    email_template: str,
    entry: ManifestEntry,
) -> tuple[str, str]:
    variables = build_email_variables(entry)
    subject = email_sender.render_template(settings.email.subject_template, variables)
    body = email_sender.render_template(email_template, variables)
    return subject, body


def _evaluate_entry_for_send(
    entry: ManifestEntry,
    *,
    project_root: Path,
    already_sent_keys: set[tuple[int, int, str, str]],
    resend_all: bool,
) -> tuple[str, str]:
    if not resend_all and entry.resume_key() in already_sent_keys:
        return EMAIL_STATUS_SENT, "Уже отправлено ранее"

    pdf_error = validate_pdf(resolve_pdf_path(project_root, entry.pdf_path))
    if pdf_error:
        return EMAIL_STATUS_SKIPPED_PDF_MISSING, pdf_error

    email_error = validate_recipient_email(entry.recipient_email)
    if email_error:
        return EMAIL_STATUS_SKIPPED_INVALID_EMAIL, email_error

    cc_bcc_error = validate_cc_bcc(entry.cc_email, entry.bcc_email)
    if cc_bcc_error:
        return EMAIL_STATUS_SKIPPED_INVALID_EMAIL, cc_bcc_error

    return EMAIL_STATUS_PENDING, ""


class EmailPipeline:
    """Run email preview, test send, or real delivery from a manifest."""

    def __init__(self, project_root: Path, settings: Settings | None = None) -> None:
        self.project_root = project_root.resolve()
        self.settings = settings or Settings(self.project_root)
        self.email_sender = EmailSender(
            self.settings.env,
            signature_text_path=self.settings.email_signature_text_path,
            signature_html_path=self.settings.email_signature_html_path,
        )

    def _load_context(
        self,
        manifest_arg: str | None,
    ) -> tuple[Path, list[ManifestEntry], EmailReport | None, set[tuple[int, int, str, str]]]:
        manifest_path = resolve_manifest_path(
            project_root=self.project_root,
            reports_dir=self.settings.reports_dir,
            manifest_arg=manifest_arg,
        )
        entries = read_manifest(manifest_path, self.project_root)
        generation_id = extract_generation_id(manifest_path)
        existing_report = read_email_report(
            email_report_path(self.settings.reports_dir, generation_id)
        )
        already_sent = sent_resume_keys(existing_report) if existing_report else set()
        return manifest_path, entries, existing_report, already_sent

    def run_preview(self, *, manifest_arg: str | None = None) -> int:
        manifest_path, entries, _, _ = self._load_context(manifest_arg)
        email_template = load_email_template(self.settings.email_template_path)
        email_run_id = make_email_run_id(extract_generation_id(manifest_path))

        report_rows: list[EmailReportRow] = []
        errors = 0
        sendable = 0

        for entry in entries:
            subject, _ = _render_subject_and_body(
                settings=self.settings,
                email_sender=self.email_sender,
                email_template=email_template,
                entry=entry,
            )
            status, error_message = _evaluate_entry_for_send(
                entry,
                project_root=self.project_root,
                already_sent_keys=set(),
                resend_all=True,
            )
            if status == EMAIL_STATUS_PENDING:
                status = EMAIL_STATUS_DRY_RUN
                sendable += 1
            else:
                errors += 1

            report_rows.append(
                _prepare_report_row(
                    entry=entry,
                    email_run_id=email_run_id,
                    subject=subject,
                    status=status,
                    error_message=error_message,
                )
            )

        report = EmailReport(
            generation_id=extract_generation_id(manifest_path),
            report_path=email_report_path(
                self.settings.reports_dir,
                extract_generation_id(manifest_path),
            ),
            rows=report_rows,
        )
        write_email_report(report)

        manifest_rel = relative_project_path(self.project_root, manifest_path)
        print()
        print("EMAIL PREVIEW")
        print(f"Manifest: {manifest_rel}")
        print(f"Писем к отправке: {sendable}")
        print(f"Ошибок: {errors}")
        print("Реальная отправка НЕ выполнялась.")
        return 0

    def _collect_sendable_entries(self, entries: list[ManifestEntry]) -> list[ManifestEntry]:
        sendable: list[ManifestEntry] = []
        for entry in entries:
            status, _ = _evaluate_entry_for_send(
                entry,
                project_root=self.project_root,
                already_sent_keys=set(),
                resend_all=True,
            )
            if status == EMAIL_STATUS_PENDING:
                sendable.append(entry)
        return sendable

    @staticmethod
    def _render_test_subject_and_body(
        *,
        settings: Settings,
        email_sender: EmailSender,
        email_template: str,
        entry: ManifestEntry,
    ) -> tuple[str, str]:
        subject, body = _render_subject_and_body(
            settings=settings,
            email_sender=email_sender,
            email_template=email_template,
            entry=entry,
        )
        subject = f"[TEST] {subject}"
        body = (
            "ТЕСТОВАЯ ОТПРАВКА.\n"
            f"Оригинальный получатель: {entry.recipient_email}\n\n"
            f"{body}"
        )
        return subject, body

    @staticmethod
    def _make_test_report_row(
        *,
        generation_id: int,
        email_run_id: str,
        entry: ManifestEntry,
        test_email: str,
        subject: str,
        status: str,
        attempts: int,
        sent_at: str,
        error_message: str,
    ) -> EmailTestReportRow:
        return EmailTestReportRow(
            generation_id=generation_id,
            email_run_id=email_run_id,
            mode=entry.mode,
            excel_row=entry.excel_row,
            bank_name=entry.bank_name,
            bank_legal_name=entry.bank_legal_name,
            out_number=entry.out_number,
            original_recipient_email=entry.recipient_email,
            test_recipient_email=test_email,
            cc_email=entry.cc_email,
            bcc_email=entry.bcc_email,
            subject=subject,
            pdf_path=entry.pdf_path,
            status=status,
            attempts=attempts,
            sent_at=sent_at,
            error_message=error_message,
        )

    def _send_one_test_email(
        self,
        *,
        entry: ManifestEntry,
        test_email: str,
        email_template: str,
        generation_id: int,
        email_run_id: str,
    ) -> EmailTestReportRow:
        subject, body = self._render_test_subject_and_body(
            settings=self.settings,
            email_sender=self.email_sender,
            email_template=email_template,
            entry=entry,
        )
        pdf_path = resolve_pdf_path(self.project_root, entry.pdf_path)

        status = EMAIL_STATUS_TEST_FAILED
        error_message = ""
        sent_at = ""
        attempts = 1

        try:
            self.email_sender.send(
                to_email=test_email,
                cc_email="",
                bcc_email="",
                subject=subject,
                body=body,
                pdf_path=pdf_path,
                bank_name_for_log=entry.bank_name,
            )
            status = EMAIL_STATUS_TEST_SENT
            sent_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(
                f"Тест отправлен: {entry.bank_name} -> {test_email} "
                f"(оригинал: {entry.recipient_email}, СВХИ {entry.out_number})"
            )
        except EmailSenderError as exc:
            error_message = str(exc)
            print(
                f"Ошибка теста: {entry.bank_name} -> {test_email}: {exc}"
            )
            logger.error(
                "Test email failed for %s (%s): %s",
                entry.bank_name,
                entry.recipient_email,
                exc,
            )

        return self._make_test_report_row(
            generation_id=generation_id,
            email_run_id=email_run_id,
            entry=entry,
            test_email=test_email,
            subject=subject,
            status=status,
            attempts=attempts,
            sent_at=sent_at,
            error_message=error_message,
        )

    def run_test_send(
        self,
        *,
        manifest_arg: str | None = None,
        send_all: bool = False,
        recipient_email: str | None = None,
    ) -> int:
        validate_smtp_settings(self.settings)
        manifest_path, entries, _, _ = self._load_context(manifest_arg)
        email_template = load_email_template(self.settings.email_template_path)
        generation_id = extract_generation_id(manifest_path)
        email_run_id = make_email_run_id(generation_id)
        delay_seconds = self.settings.app.delay_between_emails_seconds

        sendable_entries = self._collect_sendable_entries(entries)
        if not sendable_entries:
            print("Не найдено ни одной строки manifest с валидным PDF и email.")
            return 1

        targets = sendable_entries if send_all else [sendable_entries[0]]

        try:
            test_email = resolve_test_recipient_email(
                self.settings,
                override=recipient_email,
            )
        except EmailPipelineError as exc:
            print(exc)
            return 1

        if send_all:
            print(f"Тестовых писем к отправке: {len(targets)}")
            print(f"Получатель теста: {test_email}")
            if delay_seconds > 0 and len(targets) > 1:
                wait_seconds = delay_seconds * (len(targets) - 1)
                print(
                    f"Задержка между письмами: {delay_seconds} сек "
                    f"(~{wait_seconds // 60} мин {wait_seconds % 60} сек ожидания)"
                )

        test_report_path = email_test_report_path(self.settings.reports_dir, generation_id)
        report_rows: list[EmailTestReportRow] = []
        sent_count = 0
        failed_count = 0

        for index, entry in enumerate(targets):
            report_row = self._send_one_test_email(
                entry=entry,
                test_email=test_email,
                email_template=email_template,
                generation_id=generation_id,
                email_run_id=email_run_id,
            )
            report_rows.append(report_row)
            if report_row.status == EMAIL_STATUS_TEST_SENT:
                sent_count += 1
            else:
                failed_count += 1

            write_email_test_report(
                EmailTestReport(
                    generation_id=generation_id,
                    report_path=test_report_path,
                    rows=report_rows,
                )
            )

            has_more = index < len(targets) - 1
            if has_more and delay_seconds > 0:
                logger.info("Waiting %d seconds before next test email...", delay_seconds)
                time.sleep(delay_seconds)

        print()
        print("TEST EMAIL SUMMARY")
        print(f"Отправлено: {sent_count}")
        print(f"Ошибок: {failed_count}")
        print(
            "Test email report: "
            f"{relative_project_path(self.project_root, test_report_path)}"
        )
        return 0 if failed_count == 0 else 1

    def run_preview_one_mail(
        self,
        *,
        manifest_arg: str | None = None,
        recipient_email: str | None = None,
    ) -> int:
        """Fast test: one email from preview manifest to a fixed address."""
        validate_smtp_settings(self.settings)
        try:
            test_email = resolve_test_recipient_email(
                self.settings,
                override=recipient_email,
                prompt_if_missing=False,
            )
        except EmailPipelineError as exc:
            print(exc)
            return 1
        manifest_path = resolve_preview_manifest_path(
            project_root=self.project_root,
            manifest_arg=manifest_arg,
        )
        entries = read_manifest(manifest_path, self.project_root)
        email_template = load_email_template(self.settings.email_template_path)
        generation_id = extract_generation_id(manifest_path)
        email_run_id = make_email_run_id(generation_id)
        preview_reports_dir = self.project_root / "output" / "preview" / "reports"

        sendable_entries = self._collect_sendable_entries(entries)
        if not sendable_entries:
            print("Не найдено ни одной строки preview manifest с валидным PDF и email.")
            print("Сначала выполните preview.py.")
            return 1

        target_entry = sendable_entries[0]

        manifest_rel = relative_project_path(self.project_root, manifest_path)
        pdf_path = resolve_pdf_path(self.project_root, target_entry.pdf_path)
        print()
        print("PREVIEW ONE MAIL (fast test)")
        print(f"Manifest: {manifest_rel}")
        print(f"Банк: {target_entry.bank_name} (строка Excel {target_entry.excel_row})")
        print(f"PDF: {relative_project_path(self.project_root, pdf_path)}")
        print(f"Оригинальный email банка: {target_entry.recipient_email}")
        print(f"Отправка на: {test_email}")

        subject, body = _render_subject_and_body(
            settings=self.settings,
            email_sender=self.email_sender,
            email_template=email_template,
            entry=target_entry,
        )
        subject = f"[TEST] {subject}"
        body = (
            "ТЕСТ PREVIEW (fast).\n"
            f"Оригинальный получатель: {target_entry.recipient_email}\n\n"
            f"{body}"
        )

        status = EMAIL_STATUS_TEST_FAILED
        error_message = ""
        sent_at = ""
        attempts = 1

        try:
            self.email_sender.send(
                to_email=test_email,
                cc_email="",
                bcc_email="",
                subject=subject,
                body=body,
                pdf_path=pdf_path,
                bank_name_for_log=target_entry.bank_name,
            )
            status = EMAIL_STATUS_TEST_SENT
            sent_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"Письмо отправлено на {test_email}")
        except EmailSenderError as exc:
            error_message = str(exc)
            print(f"Ошибка отправки: {exc}")
            logger.error("Preview one mail failed: %s", exc)

        test_report_path = email_test_report_path(preview_reports_dir, generation_id)
        write_email_test_report(
            EmailTestReport(
                generation_id=generation_id,
                report_path=test_report_path,
                rows=[
                    self._make_test_report_row(
                        generation_id=generation_id,
                        email_run_id=email_run_id,
                        entry=target_entry,
                        test_email=test_email,
                        subject=subject,
                        status=status,
                        attempts=attempts,
                        sent_at=sent_at,
                        error_message=error_message,
                    )
                ],
            )
        )
        print(
            "Test email report: "
            f"{relative_project_path(self.project_root, test_report_path)}"
        )
        return 0 if status == EMAIL_STATUS_TEST_SENT else 1

    def run_send(
        self,
        *,
        manifest_arg: str | None = None,
        resend_all: bool = False,
        skip_confirmation: bool = False,
    ) -> int:
        validate_smtp_settings(self.settings)
        manifest_path, entries, existing_report, already_sent = self._load_context(
            manifest_arg
        )
        email_template = load_email_template(self.settings.email_template_path)
        email_run_id = make_email_run_id(extract_generation_id(manifest_path))
        delay_seconds = self.settings.app.delay_between_emails_seconds

        summary = build_preflight_summary(
            manifest_path=manifest_path,
            entries=entries,
            project_root=self.project_root,
            already_sent_keys=already_sent if not resend_all else set(),
        )
        print_preflight_summary(
            summary,
            project_root=self.project_root,
            settings=self.settings,
            delay_seconds=delay_seconds,
        )

        if summary.sendable == 0:
            print("Нет писем для отправки.")
            return 0

        if not skip_confirmation:
            answer = input(SEND_CONFIRMATION_PROMPT).strip()
            if answer != "SEND":
                print("Рассылка отменена.")
                return 0

        existing_by_key: dict[tuple[int, int, str, str], EmailReportRow] = {}
        if existing_report:
            for row in existing_report.rows:
                existing_by_key[row.resume_key()] = row

        report_rows: list[EmailReportRow] = []
        sent_count = 0
        failed_count = 0
        skipped_count = 0

        for entry in entries:
            subject, body = _render_subject_and_body(
                settings=self.settings,
                email_sender=self.email_sender,
                email_template=email_template,
                entry=entry,
            )

            if not resend_all and entry.resume_key() in already_sent:
                previous = existing_by_key.get(entry.resume_key())
                report_rows.append(
                    previous
                    or _prepare_report_row(
                        entry=entry,
                        email_run_id=email_run_id,
                        subject=subject,
                        status=EMAIL_STATUS_SENT,
                    )
                )
                continue

            status, error_message = _evaluate_entry_for_send(
                entry,
                project_root=self.project_root,
                already_sent_keys=already_sent,
                resend_all=resend_all,
            )

            if status != EMAIL_STATUS_PENDING:
                skipped_count += 1
                report_rows.append(
                    _prepare_report_row(
                        entry=entry,
                        email_run_id=email_run_id,
                        subject=subject,
                        status=status,
                        error_message=error_message,
                    )
                )
                continue

            pdf_path = resolve_pdf_path(self.project_root, entry.pdf_path)
            attempts = 1
            sent_at = ""

            try:
                self.email_sender.send(
                    to_email=entry.recipient_email,
                    cc_email=entry.cc_email,
                    bcc_email=entry.bcc_email,
                    subject=subject,
                    body=body,
                    pdf_path=pdf_path,
                    bank_name_for_log=entry.bank_name,
                )
                status = EMAIL_STATUS_SENT
                sent_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                sent_count += 1
                print(
                    f"Отправлено: {entry.bank_name} -> {entry.recipient_email} "
                    f"(СВХИ {entry.out_number})"
                )
            except EmailSenderError as exc:
                status = EMAIL_STATUS_FAILED
                error_message = str(exc)
                failed_count += 1
                print(
                    f"Ошибка: {entry.bank_name} -> {entry.recipient_email}: {exc}"
                )
                logger.error(
                    "Email failed for %s (%s): %s",
                    entry.bank_name,
                    entry.recipient_email,
                    exc,
                )

            report_rows.append(
                _prepare_report_row(
                    entry=entry,
                    email_run_id=email_run_id,
                    subject=subject,
                    status=status,
                    attempts=attempts,
                    sent_at=sent_at,
                    error_message=error_message,
                )
            )

            report = EmailReport(
                generation_id=extract_generation_id(manifest_path),
                report_path=email_report_path(
                    self.settings.reports_dir,
                    extract_generation_id(manifest_path),
                ),
                rows=report_rows,
            )
            write_email_report(report)

            has_more_sendable = False
            for future_entry in entries[entries.index(entry) + 1 :]:
                future_status, _ = _evaluate_entry_for_send(
                    future_entry,
                    project_root=self.project_root,
                    already_sent_keys=already_sent,
                    resend_all=resend_all,
                )
                if future_status == EMAIL_STATUS_PENDING and (
                    resend_all or future_entry.resume_key() not in already_sent
                ):
                    has_more_sendable = True
                    break

            if has_more_sendable and delay_seconds > 0:
                logger.info("Waiting %d seconds before next email...", delay_seconds)
                time.sleep(delay_seconds)

        final_report = EmailReport(
            generation_id=extract_generation_id(manifest_path),
            report_path=email_report_path(
                self.settings.reports_dir,
                extract_generation_id(manifest_path),
            ),
            rows=report_rows,
        )
        write_email_report(final_report)

        print()
        print("EMAIL SEND SUMMARY")
        print(f"Отправлено: {sent_count}")
        print(f"Ошибок: {failed_count}")
        print(f"Пропущено: {skipped_count}")
        print(
            "Email report: "
            f"{relative_project_path(self.project_root, final_report.report_path)}"
        )
        return 0 if failed_count == 0 else 1
