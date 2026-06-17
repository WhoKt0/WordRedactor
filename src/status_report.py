"""Status report writer (Excel)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from src.models import StatusReportRow

REPORT_COLUMNS = [
    "row_number",
    "bank_name",
    "bank_legal_name",
    "recipient_email",
    "out_number",
    "date",
    "docx_path",
    "pdf_path",
    "status",
    "error_message",
    "sent_at",
]


def write_status_report(rows: list[StatusReportRow], reports_dir: Path) -> Path:
    """Write status report to output/reports/status_YYYYMMDD_HHMMSS.xlsx."""
    reports_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = reports_dir / f"status_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "status"

    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, report_row in enumerate(rows, start=2):
        values = [
            report_row.row_number,
            report_row.bank_name,
            report_row.bank_legal_name,
            report_row.recipient_email,
            report_row.out_number,
            report_row.date,
            report_row.docx_path,
            report_row.pdf_path,
            report_row.status,
            report_row.error_message,
            report_row.sent_at,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(report_path)
    return report_path
